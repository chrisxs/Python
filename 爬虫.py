# 1.5copy
import re
import requests
from bs4 import BeautifulSoup
import os
import openpyxl
from datetime import datetime
from tqdm import tqdm

# 创建一个images文件夹，如果该文件夹不存在
if not os.path.exists('images'):
    os.mkdir('images')

# 用户输入要爬取的网站URL
url = input('请输入要爬取的网站URL：')

# 发送请求并获取响应内容
response = requests.get(url)

# 使用BeautifulSoup解析HTML内容
soup = BeautifulSoup(response.content, 'html.parser')

# 获取网页标题
title = soup.title.string

# 获取所有图片链接
img_links = []
for img in soup.find_all('img'):
    src = img.get('src')
    if src is not None:
        # 使用正则表达式判断链接是否以http或https开头
        if re.match(r'^https?://', src):
            img_links.append(src)

# 下载图片并保存到images文件夹
for i, img_url in enumerate(img_links):
    try:
        response = requests.get(img_url)
        with open(f'images/img_{i}.jpg', 'wb') as f:
            f.write(response.content)
    except Exception as e:
        print(f'下载图片 {img_url} 时发生异常: {e}')

# 获取所有链接和链接标题
links = []
for link in soup.find_all('a'):
    href = link.get('href')
    title = link.string
    if href is not None:
        links.append({'title': title, 'href': href})

# 创建工作簿和工作表
wb = openpyxl.Workbook()
ws = wb.active

# 写入标题和表头
ws['A1'] = '网页标题'
ws['B1'] = '图片链接'
ws['C1'] = '链接标题'
ws['D1'] = '链接地址'

# 写入数据
for i in range(len(img_links)):
    img_url = img_links[i]
    if img_url.startswith('http'):
        response = requests.get(img_url)
        if response.status_code == 200:
            # 获取文件名并将特殊字符替换成 '_'
            filename = os.path.basename(img_url)
            filename = re.sub(r'[\\/*?:"<>|]', '_', filename)
            # 创建images文件夹（如果不存在）
            if not os.path.exists('images'):
                os.makedirs('images')
            # 下载并保存图片到本地
            with open('images/' + filename, 'wb') as f:
                f.write(response.content)
            # 写入图片链接到Excel
            ws.cell(row=i+2, column=2, value=img_url)
        else:
            print(f"图片 {img_url} 下载失败")
    else:
        print(f"无效的图片链接：{img_url}")


for i in range(len(links)):
    if links[i]['title'] is not None:
        ws.cell(row=i+2, column=3, value=links[i]['title'])
    ws.cell(row=i+2, column=4, value=links[i]['href'])
ws['A2'] = title

# 保存工作簿
now = datetime.now()
filename = now.strftime('%Y-%m-%d %H-%M-%S') + '.xlsx'
wb.save(filename)

# 创建目录 images
if not os.path.exists('images'):
    os.mkdir('images')

# 下载所有图片并保存到本地
# 下载所有图片并保存到本地
with tqdm(total=len(img_links), desc='下载图片') as pbar:
    for i, img_url in enumerate(img_links):
        # 如果图片链接不是完整的链接，补全链接
        if not img_url.startswith('http'):
            img_url = url + \
                img_url if img_url.startswith('/') else url + '/' + img_url

        # 下载图片并保存到本地
        response = requests.get(img_url)
        if response.status_code == 200:
            with open('images/{}.jpg'.format(i+1), 'wb') as f:
                f.write(response.content)
            pbar.update(1)  # 更新进度条
        else:
            print('下载第{}张图片失败'.format(i+1))

# 在IDLE SHELL中打印结果
print('网页标题：')
print(title)
print('\n图片链接：')
for link in img_links:
    print(link)
print('\n链接标题和地址：')
for link in links:
    if link['title'] is not None:
        print(link['title'] + ': ' + link['href'])
    else:
        print(link['href'])

print(f"\n图片已保存到 {os.path.abspath('images')} 目录下")
